import sys
import csv
import logging
import pandas as pd
import numpy as np
from typing import Tuple, Dict
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QGridLayout, QWidget, QPushButton,
    QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QColorDialog, QComboBox,
    QMessageBox, QFileDialog, QCheckBox, QTabWidget,
    QMenuBar, QMenu, QAction, QDialog, QSpinBox, QFormLayout, QDialogButtonBox,
    QGroupBox, QScrollArea, QTextEdit, QSizePolicy
)
from PyQt5.QtGui import QColor, QFont
from PyQt5.QtCore import Qt, QTimer
import pyqtgraph as pg
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime
from io import StringIO
import re
from scipy.optimize import curve_fit
from scipy.signal import find_peaks

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# helper functions for fmg file export
def rgb_to_decimal(color_str):
    """Convert RGB hex color to decimal integer for FLIPR format"""
    # Remove '#' if present and convert to int
    color = int(color_str.lstrip('#'), 16)
    return color

def format_concentration(conc_str):
    """Format concentration string to numeric value"""
    if not conc_str:
        return 0
    # Extract numeric value from string like "10 µM"
    match = re.search(r'([\d.]+)', conc_str)
    return float(match.group(1)) if match else 0


# class definitions
class ParametersDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Analysis Parameters")
        self.setModal(True)

        # Create layout
        layout = QFormLayout()

        # Create spinboxes for parameters
        self.artifact_start = QSpinBox()
        self.artifact_start.setRange(0, 1000)
        self.artifact_start.setValue(18)
        self.artifact_start.setToolTip("Start frame of injection artifact")

        self.artifact_end = QSpinBox()
        self.artifact_end.setRange(0, 1000)
        self.artifact_end.setValue(30)
        self.artifact_end.setToolTip("End frame of injection artifact")

        self.baseline_frames = QSpinBox()
        self.baseline_frames.setRange(1, 100)
        self.baseline_frames.setValue(15)
        self.baseline_frames.setToolTip("Number of frames for baseline calculation")

        self.peak_start_frame = QSpinBox()
        self.peak_start_frame.setRange(0, 1000)
        self.peak_start_frame.setValue(20)
        self.peak_start_frame.setToolTip("Start frame for peak detection")

        # Add widgets to layout
        layout.addRow("Artifact Start Frame:", self.artifact_start)
        layout.addRow("Artifact End Frame:", self.artifact_end)
        layout.addRow("Baseline Frames:", self.baseline_frames)
        layout.addRow("Peak Start Frame:", self.peak_start_frame)

        # Add buttons
        buttons = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel,
            Qt.Horizontal, self
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)

        layout.addRow(buttons)
        self.setLayout(layout)


class BasePlotWindow(QMainWindow):
    """Base class for all plot windows"""
    def __init__(self, title="Plot Window", parent=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.resize(800, 600)
        self.was_visible = False

        # Create central widget and layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        self.layout = QVBoxLayout(central_widget)

        # Create plot widget with white background
        self.plot_widget = pg.PlotWidget()
        self.plot_widget.setBackground('w')
        self.plot_widget.setLabel('left', "Signal")
        self.plot_widget.setLabel('bottom', "Time (s)")
        self.plot_widget.addLegend()

        # Create control panel
        control_panel = QWidget()
        control_layout = QHBoxLayout(control_panel)

        # Add clear button
        self.clear_button = QPushButton("Clear Plot")
        self.clear_button.clicked.connect(self.clear_plot)

        # Add show grid checkbox
        self.grid_checkbox = QCheckBox("Show Grid")
        self.grid_checkbox.stateChanged.connect(self.toggle_grid)

        control_layout.addWidget(self.clear_button)
        control_layout.addWidget(self.grid_checkbox)
        control_layout.addStretch()

        # Add widgets to main layout
        self.layout.addWidget(self.plot_widget)
        self.layout.addWidget(control_panel)

        self.plot_items = {}

    def closeEvent(self, event):
        self.was_visible = False
        super().closeEvent(event)

    def clear_plot(self):
        self.plot_widget.clear()
        self.plot_items = {}
        self.plot_widget.addLegend()

    def toggle_grid(self, state):
        self.plot_widget.showGrid(x=state, y=state)

    def plot_trace(self, well: str, times, values, color='b'):
        if well in self.plot_items:
            self.plot_widget.removeItem(self.plot_items[well])
        pen = pg.mkPen(color=color, width=2)
        self.plot_items[well] = self.plot_widget.plot(times, values, pen=pen, name=well)

class RawPlotWindow(BasePlotWindow):
    def __init__(self, parent=None):
        super().__init__(title="Raw Traces", parent=parent)
        self.plot_widget.setLabel('left', "Intensity (A.U.)")

class DFFPlotWindow(BasePlotWindow):
    def __init__(self, parent=None):
        super().__init__(title="ΔF/F₀ Traces", parent=parent)
        self.plot_widget.setLabel('left', "Intensity (ΔF/F₀)")


class SummaryPlotWindow(QMainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Summary Plots")
        self.resize(800, 600)

        # Create central widget and layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # Add tab widget for different summary plots
        self.tab_widget = QTabWidget()
        main_layout.addWidget(self.tab_widget)

        # Create tabs for different plot types
        self.individual_tab = QWidget()
        self.mean_tab = QWidget()
        self.responses_tab = QWidget()
        self.auc_tab = QWidget()  # New tab for AUC
        self.time_to_peak_tab = QWidget()  # New tab for Time to Peak
        self.normalized_tab = QWidget()

        self.tab_widget.addTab(self.individual_tab, "Individual Traces")
        self.tab_widget.addTab(self.mean_tab, "Mean Traces")
        self.tab_widget.addTab(self.responses_tab, "Peak Responses")
        self.tab_widget.addTab(self.auc_tab, "Area Under Curve")
        self.tab_widget.addTab(self.time_to_peak_tab, "Time to Peak")
        self.tab_widget.addTab(self.normalized_tab, "Normalized to Ionomycin")


        # Create plot widgets for each tab
        self.individual_plot = pg.PlotWidget()
        self.individual_plot.setBackground('w')
        self.individual_plot.setLabel('left', "ΔF/F₀")
        self.individual_plot.setLabel('bottom', "Time (s)")

        self.mean_plot = pg.PlotWidget()
        self.mean_plot.setBackground('w')
        self.mean_plot.setLabel('left', "ΔF/F₀")
        self.mean_plot.setLabel('bottom', "Time (s)")

        self.responses_plot = pg.PlotWidget()
        self.responses_plot.setBackground('w')
        self.responses_plot.setLabel('left', "Peak ΔF/F₀")
        self.responses_plot.setLabel('bottom', "Group")

        self.normalized_plot = pg.PlotWidget()
        self.normalized_plot.setBackground('w')
        self.normalized_plot.setLabel('left', "Response (% Ionomycin)")
        self.normalized_plot.setLabel('bottom', "Group")

        # Create plot widgets for AUC and Time to Peak
        self.auc_plot = pg.PlotWidget()
        self.auc_plot.setBackground('w')
        self.auc_plot.setLabel('left', "Area Under Curve")
        self.auc_plot.setLabel('bottom', "Group")

        self.time_to_peak_plot = pg.PlotWidget()
        self.time_to_peak_plot.setBackground('w')
        self.time_to_peak_plot.setLabel('left', "Time to Peak (s)")
        self.time_to_peak_plot.setLabel('bottom', "Group")

        # Set up layouts for all tabs
        individual_layout = QVBoxLayout(self.individual_tab)
        individual_layout.addWidget(self.individual_plot)

        mean_layout = QVBoxLayout(self.mean_tab)
        mean_layout.addWidget(self.mean_plot)

        responses_layout = QVBoxLayout(self.responses_tab)
        responses_layout.addWidget(self.responses_plot)

        auc_layout = QVBoxLayout(self.auc_tab)
        auc_layout.addWidget(self.auc_plot)

        time_to_peak_layout = QVBoxLayout(self.time_to_peak_tab)
        time_to_peak_layout.addWidget(self.time_to_peak_plot)

        normalized_layout = QVBoxLayout(self.normalized_tab)
        normalized_layout.addWidget(self.normalized_plot)

        self.plot_items = {}  # Store plot items for reference

    def clear_plots(self):
        """Clear all plots"""
        self.individual_plot.clear()
        self.mean_plot.clear()
        self.responses_plot.clear()
        self.auc_plot.clear()
        self.time_to_peak_plot.clear()
        self.normalized_plot.clear()
        self.plot_items = {}

class DataProcessor:
    """Class to handle data processing operations"""
    @staticmethod
    def get_F0(data: pd.DataFrame, baseline_frames: int = 15) -> np.ndarray:
        """Calculate baseline (F0) as mean of first baseline_frames for each row."""
        return data.iloc[:, :baseline_frames].mean(axis=1)

    @staticmethod
    def calculate_dff(data: pd.DataFrame, F0: np.ndarray) -> pd.DataFrame:
        """Calculate ΔF/F₀"""
        return (data.div(F0, axis=0) -1)

    @staticmethod
    def calculate_peak_response(data: pd.DataFrame, start_frame: int = None) -> pd.Series:
        """Calculate peak response"""
        if start_frame is not None:
            data = data.iloc[:, start_frame:]
        return data.max(axis=1)

    @staticmethod
    def calculate_auc(data: pd.DataFrame, time_points: np.ndarray) -> pd.Series:
        """Calculate area under the curve using trapezoidal integration"""
        return pd.Series({
            well: np.trapz(y=data.loc[well], x=time_points)
            for well, row in data.iterrows()
        })

class PeakAnalyzer:
    """Class for peak detection and fitting analysis"""

    @staticmethod
    def peak_function(x, amplitude, center, sigma, tau_rise, tau_decay):
        """Define peak shape function - asymmetric gaussian with rise and decay"""
        y = np.zeros_like(x)
        for i, t in enumerate(x):
            if t <= center:
                # Rising phase
                y[i] = amplitude * (1 - np.exp(-(t - (center - 5*tau_rise))/tau_rise))
            else:
                # Decay phase
                y[i] = amplitude * np.exp(-(t - center)/tau_decay)
        return y

    def analyze_trace(self, times, values):
        """Analyze a single trace"""
        try:
            # Find initial peak parameters
            peaks, properties = find_peaks(values, prominence=0.2*np.max(values))
            if len(peaks) == 0:
                return None

            peak_idx = peaks[np.argmax(values[peaks])]
            peak_value = values[peak_idx]
            peak_time = times[peak_idx]

            # Initial parameter guesses
            p0 = [
                peak_value,  # amplitude
                peak_time,   # center
                2.0,        # sigma
                2.0,        # tau_rise
                5.0         # tau_decay
            ]

            # Fit curve
            popt, _ = curve_fit(self.peak_function, times, values, p0=p0)

            # Calculate metrics
            fitted_curve = self.peak_function(times, *popt)
            auc = np.trapz(y=fitted_curve, x=times)

            # Find FWHM
            half_max = popt[0] / 2
            rise_time = self.find_rise_time(times, fitted_curve, popt[1])
            fwhm = self.find_fwhm(times, fitted_curve)

            return {
                'amplitude': popt[0],
                'peak_time': popt[1],
                'rise_time': rise_time,
                'tau_decay': popt[4],
                'fwhm': fwhm,
                'auc': auc,
                'fitted_curve': fitted_curve
            }

        except Exception as e:
            logger.error(f"Peak fitting failed: {str(e)}")
            return None

    @staticmethod
    def find_rise_time(times, values, peak_time):
        """Calculate 10-90% rise time"""
        peak_idx = np.argmin(np.abs(times - peak_time))
        max_val = values[peak_idx]
        t10_idx = np.argmin(np.abs(values[:peak_idx] - 0.1*max_val))
        t90_idx = np.argmin(np.abs(values[:peak_idx] - 0.9*max_val))
        return times[t90_idx] - times[t10_idx]

    @staticmethod
    def find_fwhm(times, values):
        """Calculate Full Width at Half Maximum"""
        max_val = np.max(values)
        half_max = max_val / 2
        above_half = values >= half_max
        regions = np.diff(above_half.astype(int))
        rising = np.where(regions == 1)[0]
        falling = np.where(regions == -1)[0]
        if len(rising) > 0 and len(falling) > 0:
            return times[falling[0]] - times[rising[0]]
        return None

class DraggableWellButton(QPushButton):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.main_window = parent
        self.setMouseTracking(True)
        # Make sure button stays square
        self.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)  # Changed to Fixed
        self.setMinimumSize(90, 90)  # Set minimum size

    def resizeEvent(self, event):
        """Keep the button square by using the minimum of width and height"""
        super().resizeEvent(event)
        size = min(self.width(), self.height())
        self.setFixedSize(size, size)

    def mousePressEvent(self, event):
        """Handle mouse press for selection"""
        if event.button() == Qt.LeftButton:
            well_index = self.property("well_index")
            if well_index is not None:
                self.main_window.toggle_well_selection(int(well_index))

    def enterEvent(self, event):
        """Handle mouse enter for shift-selection"""
        modifiers = QApplication.keyboardModifiers()
        if modifiers == Qt.ShiftModifier:
            well_index = self.property("well_index")
            print(f"\nShift + Mouse Enter on Well {well_index}")
            current_index = int(well_index)

            if hasattr(self.main_window, 'last_selected') and self.main_window.last_selected is not None:
                self.main_window.add_rectangle_to_selection(self.main_window.last_selected, current_index)



class WellPlateLabeler(QMainWindow):
    """Main application window"""
    def __init__(self):
        super().__init__()
        self.setWindowTitle("FLIPR Analysis")

        self.last_selected = None

        self.font_size = 11  # Default font size

        # Set window size and font
        self.resize(800, 600)
        self.default_font = QFont()
        self.default_font.setPointSize(12)
        QApplication.setFont(self.default_font)

        # Create file info and status bar group
        info_widget = QWidget()
        info_layout = QHBoxLayout()
        info_layout.setSpacing(2)
        info_layout.setContentsMargins(5, 0, 5, 0)

        # Add file label and display
        file_label = QLabel("Current File:")
        self.file_display = QLineEdit()
        self.file_display.setReadOnly(True)
        self.file_display.setPlaceholderText("No file loaded")

        # Add status message display
        self.status_display = QLineEdit()
        self.status_display.setReadOnly(True)
        self.status_display.setPlaceholderText("Ready")

        # Add widgets to info layout
        info_layout.addWidget(file_label)
        info_layout.addWidget(self.file_display, stretch=1)  # Give file display more space
        info_layout.addWidget(self.status_display, stretch=2)  # Give status display even more space
        info_widget.setLayout(info_layout)

        # Create main layout
        main_widget = QWidget()
        main_layout = QVBoxLayout()
        main_layout.setSpacing(2)
        main_layout.setContentsMargins(5, 5, 5, 5)

        # Add info widget at top
        main_layout.addWidget(info_widget)

        # Add tab widget
        self.tab_widget = QTabWidget()
        main_layout.addWidget(self.tab_widget)

        main_widget.setLayout(main_layout)
        self.setCentralWidget(main_widget)

        # Initialize rest of the GUI
        self.init_data()
        self.setup_plate_tab()
        self.setup_analysis_tab()
        self.create_menus()



    def init_data(self):
        """Initialize all data attributes"""
        self.analysis_params = {
            'artifact_start': 18,
            'artifact_end': 30,
            'baseline_frames': 15,
            'peak_start_frame': 20
        }

        self.remove_artifact = False
        self.normalize_to_ionomycin = False
        self.raw_plot_window = RawPlotWindow()
        self.dff_plot_window = DFFPlotWindow()
        self.summary_plot_window = SummaryPlotWindow()
        self.processor = DataProcessor()
        self.raw_data = None
        self.dff_data = None
        self.processed_time_points = None

        self.default_colors = [
            '#1f77b4', '#ff7f0e', '#2ca02c', '#d62728',
            '#9467bd', '#8c564b', '#e377c2', '#7f7f7f',
        ]

        self.well_data = [
            {
                "well_id": "",
                "label": "",
                "concentration": "",
                "sample_id": "",
                "color": self.default_colors[i % len(self.default_colors)]
            }
            for i in range(96)
        ]

        self.selected_wells = set()
        self.current_color = QColor(self.default_colors[0])

        self.selection_state = {
            'rows': set(),  # Track selected rows
            'cols': set(),  # Track selected columns
            'wells': set(),  # Track individually selected wells
            'all_selected': False  # Track if all wells are selected
        }

    def clear_selection_state(self):
        """Reset all selection states"""
        self.selection_state = {
            'rows': set(),
            'cols': set(),
            'wells': set(),
            'all_selected': False
        }
        self.last_selected = None

    def create_compact_input_fields(self):
        """Create more compact input fields layout"""
        input_group = QGroupBox("Well Properties")
        input_layout = QGridLayout()
        input_layout.setSpacing(2)
        input_layout.setContentsMargins(5, 5, 5, 5)

        # Make input fields more compact
        self.label_input = QLineEdit()
        self.label_input.setPlaceholderText("Agonist label")
        self.label_checkbox = QCheckBox()
        self.label_checkbox.setChecked(True)

        self.starting_conc_input = QLineEdit()
        self.starting_conc_input.setPlaceholderText("Conc (µM)")
        self.concentration_checkbox = QCheckBox()
        self.concentration_checkbox.setChecked(True)

        self.sample_id_input = QLineEdit()
        self.sample_id_input.setPlaceholderText("Sample ID")
        self.sample_id_checkbox = QCheckBox()
        self.sample_id_checkbox.setChecked(True)

        self.color_button = QPushButton()
        self.color_button.setMaximumWidth(50)
        self.color_button.clicked.connect(self.select_color)
        self.color_checkbox = QCheckBox()
        self.color_checkbox.setChecked(True)

        # Add to layout with minimal spacing
        labels = ["Agonist:", "Conc:", "ID:", "Color:"]
        widgets = [
            (self.label_input, self.label_checkbox),
            (self.starting_conc_input, self.concentration_checkbox),
            (self.sample_id_input, self.sample_id_checkbox),
            (self.color_button, self.color_checkbox)
        ]

        for i, (label, (widget, checkbox)) in enumerate(zip(labels, widgets)):
            input_layout.addWidget(QLabel(label), i, 0)
            input_layout.addWidget(widget, i, 1)
            input_layout.addWidget(checkbox, i, 2)

        input_group.setLayout(input_layout)
        return input_group

    def show_status(self, message: str, duration: int = 0):
        """Display a status message. Duration in ms (0 = permanent)"""
        self.status_display.setText(message)
        if duration > 0:
            QTimer.singleShot(duration, lambda: self.status_display.setText("Ready"))




    def create_compact_action_buttons(self):
        """Create action buttons in a more compact layout"""
        action_layout = QGridLayout()
        action_layout.setSpacing(2)

        buttons = [
            ("Apply Label", self.apply_label),
            ("Clear Selection", self.clear_selection),
            ("Save Layout", self.save_layout),
            ("Load Layout", self.load_layout),
            ("Load Data", self.open_file_dialog)
        ]

        for i, (text, callback) in enumerate(buttons):
            btn = QPushButton(text)
            btn.clicked.connect(callback)
            btn.setMaximumWidth(100)
            row = 0
            col = i
            action_layout.addWidget(btn, row, col)

        return action_layout

    def create_compact_well_plate_grid(self):
        """Create well grid with larger buttons"""
        plate_widget = QWidget()
        plate_layout = QGridLayout()
        plate_layout.setSpacing(2)

        rows = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
        cols = range(1, 13)
        self.wells = []

        # Create the select-all button
        select_all_btn = QPushButton("✓")
        select_all_btn.setStyleSheet("""
            QPushButton {
                background-color: lightgray;
                padding: 2px;
                min-width: 30px;
                min-height: 30px;
                font-size: 10pt;
            }
        """)
        select_all_btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        select_all_btn.clicked.connect(self.toggle_all_selection)
        plate_layout.addWidget(select_all_btn, 0, 0)

        # Column headers - wide but short
        for j, col in enumerate(cols):
            col_btn = QPushButton(str(col))
            col_btn.setStyleSheet("""
                QPushButton {
                    background-color: lightgray;
                    padding: 2px;
                    min-width: 90px;
                    min-height: 30px;
                    font-size: 10pt;
                }
            """)
            col_btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
            col_btn.clicked.connect(lambda checked, col=j: self.toggle_column_selection(col))
            plate_layout.addWidget(col_btn, 0, j + 1)

        # Row headers - narrow but tall
        for i, row in enumerate(rows):
            row_btn = QPushButton(row)
            row_btn.setStyleSheet("""
                QPushButton {
                    background-color: lightgray;
                    padding: 2px;
                    min-width: 30px;
                    min-height: 90px;
                    font-size: 10pt;
                }
            """)
            row_btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
            row_btn.clicked.connect(lambda checked, row=i: self.toggle_row_selection(row))
            plate_layout.addWidget(row_btn, i + 1, 0)

            # Wells
            for j, col in enumerate(cols):
                well_index = i * 12 + j
                well_id = f"{row}{col}"
                button = DraggableWellButton(self)
                button.setProperty("well_index", well_index)
                button.setStyleSheet(f"""
                    QPushButton {{
                        background-color: white;
                        padding: 2px;
                        min-width: 90px;
                        min-height: 90px;
                        font-size: {self.font_size}pt;
                        text-align: center;
                    }}
                """)
                self.wells.append(button)
                self.well_data[well_index]["well_id"] = well_id
                self.update_well_button_text(well_index)
                plate_layout.addWidget(button, i + 1, j + 1)

        # Remove stretch factors - we don't want anything to stretch
        for i in range(13):
            plate_layout.setColumnStretch(i, 0)
        for i in range(9):
            plate_layout.setRowStretch(i, 0)

        # Add spacer items to center the grid
        plate_layout.setRowStretch(9, 1)
        plate_layout.setColumnStretch(13, 1)

        plate_widget.setLayout(plate_layout)
        return plate_widget


    def update_well_appearances(self):
        """Update the visual appearance of all wells based on selection state"""
        for idx in range(96):
            well_data = self.well_data[idx]
            is_selected = idx in self.selected_wells

            color = 'lightblue' if is_selected else well_data['color']

            self.wells[idx].setStyleSheet(f"""
                QPushButton {{
                    background-color: {color};
                    padding: 2px;
                    font-size: 12pt;
                    color: black;
                }}
            """)

    def update_well_button_text(self, index):
        """Update well button text with better formatting"""
        data = self.well_data[index]
        well_id = data["well_id"]

        # Create multi-line text with proper spacing
        text_parts = [well_id]
        if data["label"]:
            text_parts.append(data["label"])
        if data["concentration"]:
            conc = data["concentration"].replace(" µM", "µ")
            text_parts.append(conc)
        if data["sample_id"]:
            text_parts.append(data["sample_id"])

        button_text = "\n".join(text_parts)

        # Update button text and style
        self.wells[index].setText(button_text)
        self.wells[index].setStyleSheet(f"""
            QPushButton {{
                background-color: {data['color']};
                padding: 5px;
                min-width: 80px;
                min-height: 80px;
                font-size: 12pt;
                text-align: center;
                qproperty-alignment: AlignCenter;
            }}
        """)

    def select_color(self):
        """Open color dialog and set current color"""
        color = QColorDialog.getColor()
        if color.isValid():
            self.current_color = color
            self.color_button.setStyleSheet(f"background-color: {color.name()}")


    def setup_plate_tab(self):
        """Set up the plate layout tab with side panel and larger grid"""
        plate_tab = QWidget()
        plate_layout = QHBoxLayout()  # Change to horizontal layout
        plate_layout.setSpacing(10)

        # Create left sidebar for controls
        sidebar = self.create_sidebar_controls()
        plate_layout.addWidget(sidebar)

        # Create main panel with just the grid
        main_panel = self.create_main_panel()
        plate_layout.addWidget(main_panel)

        # Set stretch factors to give more space to the grid
        plate_layout.setStretchFactor(sidebar, 0)  # Don't stretch sidebar
        plate_layout.setStretchFactor(main_panel, 1)  # Allow main panel to stretch

        plate_tab.setLayout(plate_layout)
        self.tab_widget.addTab(plate_tab, "Plate Layout")


    def create_sidebar_controls(self):
        """Create vertical sidebar with controls"""
        sidebar = QWidget()
        sidebar.setMinimumWidth(200)  # Smaller minimum width
        sidebar.setMaximumWidth(250)  # Smaller maximum width
        sidebar_layout = QVBoxLayout()
        sidebar_layout.setSpacing(5)  # Reduced spacing

        # Add title
        title = QLabel("Plate Controls")
        title.setStyleSheet("font-size: 12pt; font-weight: bold;")
        sidebar_layout.addWidget(title)

        # Mode selection
        mode_group = QGroupBox("Selection Mode")
        mode_layout = QVBoxLayout()
        mode_layout.setSpacing(2)  # Reduced spacing
        self.mode_selector = QComboBox()
        self.mode_selector.addItems(["Simple Label", "Log10 Series", "Clear Wells"])
        mode_layout.addWidget(self.mode_selector)
        mode_group.setLayout(mode_layout)
        sidebar_layout.addWidget(mode_group)

        # Well properties inputs
        props_group = QGroupBox("Well Properties")
        props_layout = QGridLayout()
        props_layout.setSpacing(2)  # Reduced spacing

        # Make input fields smaller
        input_height = 25  # Reduced height for input fields

        # Label input
        props_layout.addWidget(QLabel("Agonist:"), 0, 0)
        self.label_input = QLineEdit()
        self.label_input.setFixedHeight(input_height)
        props_layout.addWidget(self.label_input, 0, 1)
        self.label_checkbox = QCheckBox()
        self.label_checkbox.setChecked(True)
        props_layout.addWidget(self.label_checkbox, 0, 2)

        # Concentration input
        props_layout.addWidget(QLabel("Conc (µM):"), 1, 0)
        self.starting_conc_input = QLineEdit()
        self.starting_conc_input.setFixedHeight(input_height)
        props_layout.addWidget(self.starting_conc_input, 1, 1)
        self.concentration_checkbox = QCheckBox()
        self.concentration_checkbox.setChecked(True)
        props_layout.addWidget(self.concentration_checkbox, 1, 2)

        # Sample ID input
        props_layout.addWidget(QLabel("Sample ID:"), 2, 0)
        self.sample_id_input = QLineEdit()
        self.sample_id_input.setFixedHeight(input_height)
        props_layout.addWidget(self.sample_id_input, 2, 1)
        self.sample_id_checkbox = QCheckBox()
        self.sample_id_checkbox.setChecked(True)
        props_layout.addWidget(self.sample_id_checkbox, 2, 2)

        # Color selection
        props_layout.addWidget(QLabel("Color:"), 3, 0)
        color_layout = QHBoxLayout()
        self.color_button = QPushButton()
        self.color_button.setFixedHeight(input_height)
        self.color_button.setMaximumWidth(50)
        self.color_button.clicked.connect(self.select_color)
        color_layout.addWidget(self.color_button)
        self.color_checkbox = QCheckBox()
        self.color_checkbox.setChecked(True)
        color_layout.addWidget(self.color_checkbox)
        props_layout.addLayout(color_layout, 3, 1, 1, 2)

        props_group.setLayout(props_layout)
        sidebar_layout.addWidget(props_group)

        # Action buttons
        actions_group = QGroupBox("Actions")
        actions_layout = QVBoxLayout()
        actions_layout.setSpacing(2)  # Reduced spacing

        buttons = [
            ("Apply Label", self.apply_label),
            ("Clear Selection", self.clear_selection),
            ("Save Layout", self.save_layout),
            ("Load Layout", self.load_layout),
            ("Load Data", self.open_file_dialog)
        ]

        for text, callback in buttons:
            btn = QPushButton(text)
            btn.clicked.connect(callback)
            btn.setFixedHeight(30)  # Smaller button height
            actions_layout.addWidget(btn)

        actions_group.setLayout(actions_layout)
        sidebar_layout.addWidget(actions_group)

        # Add font size control group
        font_group = QGroupBox("Button Text Size")
        font_layout = QHBoxLayout()

        # Add spinbox for font size
        self.font_size_spin = QSpinBox()
        self.font_size_spin.setRange(6, 20)  # Reasonable font size range
        self.font_size_spin.setValue(self.font_size)
        self.font_size_spin.setSuffix("pt")
        self.font_size_spin.valueChanged.connect(self.update_font_size)

        # Add label
        font_layout.addWidget(QLabel("Font Size:"))
        font_layout.addWidget(self.font_size_spin)

        font_group.setLayout(font_layout)
        sidebar_layout.addWidget(font_group)

        # Add stretch at the bottom
        sidebar_layout.addStretch()

        sidebar.setLayout(sidebar_layout)
        return sidebar

    def update_font_size(self, new_size):
        """Update font size for all well buttons"""
        self.font_size = new_size
        # Update all well buttons
        for i in range(len(self.wells)):
            self.update_well_button_text(i)

    def update_well_button_text(self, index):
        """Update well button text with dynamic font size"""
        data = self.well_data[index]
        well_id = data["well_id"]

        # Create multi-line text with proper spacing
        text_parts = [well_id]
        if data["label"]:
            text_parts.append(data["label"])
        if data["concentration"]:
            conc = data["concentration"].replace(" µM", "µ")
            text_parts.append(conc)
        if data["sample_id"]:
            text_parts.append(data["sample_id"])

        button_text = "\n".join(text_parts)

        # Update button text and style with dynamic font size
        self.wells[index].setText(button_text)
        self.wells[index].setStyleSheet(f"""
            QPushButton {{
                background-color: {data['color']};
                padding: 2px;
                min-width: 90px;
                min-height: 90px;
                font-size: {self.font_size}pt;
                text-align: center;
                qproperty-alignment: AlignCenter;
            }}
        """)

    def create_main_panel(self):
        """Create main panel with just the well grid"""
        main_panel = QWidget()
        main_layout = QVBoxLayout()
        main_layout.setSpacing(2)

        # Add well plate grid
        grid = self.create_compact_well_plate_grid()
        main_layout.addWidget(grid)

        main_panel.setLayout(main_layout)
        return main_panel

    def setup_analysis_tab(self):
        """Set up the analysis tab"""
        analysis_tab = QWidget()
        analysis_layout = QVBoxLayout()

        # Plot controls
        plot_group = QGroupBox("Plot Controls")
        plot_layout = QVBoxLayout()

        # Plot type selection
        self.setup_plot_controls(plot_layout)

        # Analysis parameters
        params_group = QWidget()
        params_layout = QHBoxLayout()

        # Add checkboxes
        self.artifact_checkbox = QCheckBox("Remove Injection Artifact")
        self.artifact_checkbox.setChecked(self.remove_artifact)
        self.artifact_checkbox.stateChanged.connect(self.toggle_artifact_removal)
        params_layout.addWidget(self.artifact_checkbox)

        self.ionomycin_checkbox = QCheckBox("Normalize to Ionomycin")
        self.ionomycin_checkbox.setChecked(self.normalize_to_ionomycin)
        self.ionomycin_checkbox.stateChanged.connect(self.toggle_ionomycin_normalization)
        params_layout.addWidget(self.ionomycin_checkbox)

        params_group.setLayout(params_layout)
        plot_layout.addWidget(params_group)

        plot_group.setLayout(plot_layout)
        analysis_layout.addWidget(plot_group)

        # Add results display area (placeholder for now)
        results_group = QGroupBox("Analysis Results")
        results_layout = QVBoxLayout()

        # Add text area for results
        self.results_text = QTextEdit()
        self.results_text.setReadOnly(True)
        self.results_text.setMinimumHeight(200)
        results_layout.addWidget(self.results_text)

        # Add export button
        export_button = QPushButton("Export Results")
        export_button.clicked.connect(self.export_results)
        results_layout.addWidget(export_button)

        results_group.setLayout(results_layout)
        analysis_layout.addWidget(results_group)

        analysis_tab.setLayout(analysis_layout)
        self.tab_widget.addTab(analysis_tab, "Analysis")

    def update_results_text(self):
        """Update the results text display with summary statistics"""
        if self.dff_data is None:
            self.results_text.setText("No data loaded")
            return

        # Create string buffer for results
        buffer = StringIO()
        buffer.write("Analysis Results Summary\n")
        buffer.write("=" * 50 + "\n\n")

        # Get grouped data
        grouped_data = self.group_data_by_metadata()

        # Calculate and display statistics for each group
        for group_name, well_ids in grouped_data.items():
            group_data = self.dff_data.loc[well_ids]
            group_auc = self.auc_data[well_ids]

            # Calculate peak responses
            peaks = group_data.max(axis=1)
            peak_mean = peaks.mean()
            peak_sem = peaks.std() / np.sqrt(len(peaks))

            # Calculate time to peak
            peak_times = group_data.idxmax(axis=1).astype(float)
            time_to_peak_mean = peak_times.mean()
            time_to_peak_sem = peak_times.std() / np.sqrt(len(peak_times))

            # Calculate AUC statistics
            auc_mean = group_auc.mean()
            auc_sem = group_auc.std() / np.sqrt(len(group_auc))


            # Write group statistics
            buffer.write(f"Group: {group_name}\n")
            buffer.write(f"Number of wells: {len(well_ids)}\n")
            buffer.write(f"Peak ΔF/F₀: {peak_mean:.2f} ± {peak_sem:.2f} \n")
            buffer.write(f"Time to peak: {time_to_peak_mean:.2f} ± {time_to_peak_sem:.2f} s\n")
            buffer.write(f"Area Under Curve: {auc_mean:.2f} ± {auc_sem:.2f}\n")


            # Add ionomycin normalization if enabled
            if self.normalize_to_ionomycin:
                ionomycin_responses = self.get_ionomycin_responses()
                if ionomycin_responses:
                    normalized_peaks = []
                    for well_id in well_ids:
                        well_idx = next(idx for idx in range(96) if self.well_data[idx]["well_id"] == well_id)
                        sample_id = self.well_data[well_idx].get("sample_id", "default")
                        ionomycin_response = ionomycin_responses.get(sample_id)
                        if ionomycin_response:
                            peak = group_data.loc[well_id].max()
                            normalized_peaks.append((peak / ionomycin_response) * 100)

                    if normalized_peaks:
                        norm_mean = np.mean(normalized_peaks)
                        norm_sem = np.std(normalized_peaks) / np.sqrt(len(normalized_peaks))
                        buffer.write(f"Normalized response: {norm_mean:.2f} ± {norm_sem:.2f} % of ionomycin\n")

            buffer.write("\n")

        # Update text display
        self.results_text.setText(buffer.getvalue())

    def export_results(self):
        """Export results to Excel workbook"""
        if self.dff_data is None:
            self.show_status("No data to export", 3000)
            QMessageBox.warning(self, "Warning", "No data to export")
            return

        # Get save file name
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export Results", "",
            "Excel Files (*.xlsx);;All Files (*)",
            options=options
        )

        if not file_path:
            return

        try:
            self.show_status("Exporting results...")
            wb = Workbook()

            # Create all sheets
            self.create_summary_sheet(wb)
            self.create_traces_sheet(wb, "Individual_Traces", self.dff_data)
            self.create_mean_traces_sheet(wb)
            self.create_peak_responses_sheet(wb)
            self.create_analysis_metrics_sheet(wb)  # New detailed metrics sheet

            if self.normalize_to_ionomycin:
                self.create_normalized_sheet(wb)

            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            # Save workbook
            wb.save(file_path)
            self.show_status("Results exported successfully", 3000)
            QMessageBox.information(self, "Success", "Results exported successfully")

        except Exception as e:
            error_msg = f"Failed to export results: {str(e)}"
            self.show_status(error_msg, 5000)
            QMessageBox.critical(self, "Error", error_msg)

    def create_summary_sheet(self, wb):
        """Create summary sheet with statistics, concentrations, and baseline values"""
        ws = wb.create_sheet("Summary")

        # Add headers - including raw baseline stats
        headers = [
            "Group", "Agonist", "Cell ID", "Concentration (µM)", "Wells",
            "Raw Baseline (mean)", "Raw Baseline (SEM)",
            "Baseline ΔF/F₀ (mean)", "Baseline ΔF/F₀ (SEM)",
            "Peak ΔF/F₀ (mean)", "Peak ΔF/F₀ (SEM)",
            "Time to Peak (s)", "Time to Peak SEM",
            "AUC (mean)", "AUC (SEM)"
        ]
        if self.normalize_to_ionomycin:
            headers.extend(["Norm. Response (%)", "Norm. Response SEM"])

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        # Add data
        grouped_data = self.group_data_by_metadata()
        row = 2

        for group_name, well_ids in grouped_data.items():
            group_data = self.dff_data.loc[well_ids]
            raw_group_data = self.raw_data.loc[well_ids]

            # Calculate statistics
            peaks = group_data.max(axis=1)
            peak_times = group_data.idxmax(axis=1).astype(float)
            group_auc = self.auc_data[well_ids]

            # Calculate baseline values
            baseline_data = group_data.iloc[:, :self.analysis_params['baseline_frames']].mean(axis=1)
            baseline_mean = baseline_data.mean()
            baseline_sem = baseline_data.std() / np.sqrt(len(baseline_data))

            # Calculate raw baseline values
            raw_baseline_data = raw_group_data.iloc[:, :self.analysis_params['baseline_frames']].mean(axis=1)
            raw_baseline_mean = raw_baseline_data.mean()
            raw_baseline_sem = raw_baseline_data.std() / np.sqrt(len(raw_baseline_data))

            # Extract metadata from group name
            agonist = ""
            cell_id = ""
            concentration = ""
            if "|" in group_name:
                parts = [part.strip() for part in group_name.split("|")]
                if parts:
                    agonist = parts[0]
                for part in parts[1:]:
                    if "µM" in part:
                        concentration = part.replace(" µM", "")
                    elif not cell_id:
                        cell_id = part

            # Write data to cells with rounding
            current_col = 1
            ws.cell(row=row, column=current_col, value=group_name); current_col += 1
            ws.cell(row=row, column=current_col, value=agonist); current_col += 1
            ws.cell(row=row, column=current_col, value=cell_id); current_col += 1
            ws.cell(row=row, column=current_col, value=concentration); current_col += 1
            ws.cell(row=row, column=current_col, value=len(well_ids)); current_col += 1
            ws.cell(row=row, column=current_col, value=round(float(raw_baseline_mean), 3)); current_col += 1
            ws.cell(row=row, column=current_col, value=round(float(raw_baseline_sem), 3)); current_col += 1
            ws.cell(row=row, column=current_col, value=round(float(baseline_mean), 3)); current_col += 1
            ws.cell(row=row, column=current_col, value=round(float(baseline_sem), 3)); current_col += 1
            ws.cell(row=row, column=current_col, value=round(float(peaks.mean()), 3)); current_col += 1
            ws.cell(row=row, column=current_col, value=round(float(peaks.std() / np.sqrt(len(peaks))), 3)); current_col += 1
            ws.cell(row=row, column=current_col, value=round(float(peak_times.mean()), 3)); current_col += 1
            ws.cell(row=row, column=current_col, value=round(float(peak_times.std() / np.sqrt(len(peak_times))), 3)); current_col += 1
            ws.cell(row=row, column=current_col, value=round(float(group_auc.mean()), 3)); current_col += 1
            ws.cell(row=row, column=current_col, value=round(float(group_auc.std() / np.sqrt(len(group_auc))), 3)); current_col += 1

            if self.normalize_to_ionomycin:
                normalized_data = self.calculate_normalized_responses(group_name, well_ids)
                if normalized_data:
                    ws.cell(row=row, column=current_col, value=round(normalized_data['mean'], 3)); current_col += 1
                    ws.cell(row=row, column=current_col, value=round(normalized_data['sem'], 3))

            row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width



    def create_traces_sheet(self, wb, sheet_name, data):
        """Create sheet with trace data including concentrations"""
        ws = wb.create_sheet(sheet_name)

        # Add headers
        ws.cell(row=1, column=1, value="Well ID")
        ws.cell(row=1, column=2, value="Group")
        ws.cell(row=1, column=3, value="Concentration (µM)")
        for col, time in enumerate(self.processed_time_points, 4):
            ws.cell(row=1, column=col, value=float(time))

        # Add data
        row = 2
        grouped_data = self.group_data_by_metadata()
        for group_name, well_ids in grouped_data.items():
            for well_id in well_ids:
                # Get concentration for this well
                well_idx = next(idx for idx in range(96) if self.well_data[idx]["well_id"] == well_id)
                concentration = self.well_data[well_idx].get("concentration", "").replace(" µM", "")

                ws.cell(row=row, column=1, value=well_id)
                ws.cell(row=row, column=2, value=group_name)
                ws.cell(row=row, column=3, value=concentration)
                for col, value in enumerate(data.loc[well_id], 4):
                    ws.cell(row=row, column=col, value=float(value))
                row += 1

    def create_mean_traces_sheet(self, wb):
        """Create sheet with mean traces including concentrations"""
        ws = wb.create_sheet("Mean_Traces")

        # Add headers
        ws.cell(row=1, column=1, value="Group")
        ws.cell(row=1, column=2, value="Concentration (µM)")
        ws.cell(row=1, column=3, value="Time (s)")
        ws.cell(row=1, column=4, value="Mean ΔF/F₀")
        ws.cell(row=1, column=5, value="SEM")

        # Add data
        row = 2
        grouped_data = self.group_data_by_metadata()

        for group_name, well_ids in grouped_data.items():
            # Extract concentration if present
            concentration = ""
            if "|" in group_name:
                parts = group_name.split("|")
                for part in parts:
                    if "µM" in part:
                        concentration = part.strip().replace(" µM", "")

            group_data = self.dff_data.loc[well_ids]
            mean_trace = group_data.mean()
            sem_trace = group_data.sem()

            for t, (mean, sem) in enumerate(zip(mean_trace, sem_trace)):
                ws.cell(row=row, column=1, value=group_name)
                ws.cell(row=row, column=2, value=concentration)
                ws.cell(row=row, column=3, value=float(self.processed_time_points[t]))
                ws.cell(row=row, column=4, value=float(mean))
                ws.cell(row=row, column=5, value=float(sem))
                row += 1

            # Add blank row between groups
            row += 1

    def create_peak_responses_sheet(self, wb):
        """Create sheet with peak responses including raw and normalized baselines"""
        ws = wb.create_sheet("Peak_Responses")

        # Add headers
        headers = [
            "Group", "Well ID", "Concentration (µM)",
            "Raw Baseline", "Baseline ΔF/F₀", "Peak ΔF/F₀",
            "Time to Peak (s)", "AUC"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        # Add data
        row = 2
        grouped_data = self.group_data_by_metadata()

        for group_name, well_ids in grouped_data.items():
            for well_id in well_ids:
                # Get concentration for this well
                well_idx = next(idx for idx in range(96) if self.well_data[idx]["well_id"] == well_id)
                concentration = self.well_data[well_idx].get("concentration", "").replace(" µM", "")

                # Get trace data
                trace = self.dff_data.loc[well_id]
                raw_trace = self.raw_data.loc[well_id]

                # Calculate values
                raw_baseline = raw_trace.iloc[:self.analysis_params['baseline_frames']].mean()
                baseline = trace.iloc[:self.analysis_params['baseline_frames']].mean()
                peak = trace.max()
                peak_time = float(trace.idxmax())
                auc = self.auc_data[well_id]

                # Write data with rounding
                ws.cell(row=row, column=1, value=group_name)
                ws.cell(row=row, column=2, value=well_id)
                ws.cell(row=row, column=3, value=concentration)
                ws.cell(row=row, column=4, value=round(float(raw_baseline), 3))
                ws.cell(row=row, column=5, value=round(float(baseline), 3))
                ws.cell(row=row, column=6, value=round(float(peak), 3))
                ws.cell(row=row, column=7, value=round(peak_time, 3))
                ws.cell(row=row, column=8, value=round(float(auc), 3))
                row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

    def create_normalized_sheet(self, wb):
        """Create sheet with ionomycin-normalized data including concentrations"""
        if not self.normalize_to_ionomycin:
            return

        ws = wb.create_sheet("Ionomycin_Normalized")

        # Add headers
        headers = ["Group", "Well ID", "Concentration (µM)", "Normalized Response (%)",
                  "Sample ID", "Ionomycin Response"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Add data
        row = 2
        grouped_data = self.group_data_by_metadata()
        ionomycin_responses = self.get_ionomycin_responses()

        for group_name, well_ids in grouped_data.items():
            if group_name == "Ionomycin":
                continue

            for well_id in well_ids:
                well_idx = next(idx for idx in range(96) if self.well_data[idx]["well_id"] == well_id)
                concentration = self.well_data[well_idx].get("concentration", "").replace(" µM", "")
                sample_id = self.well_data[well_idx].get("sample_id", "default")
                ionomycin_response = ionomycin_responses.get(sample_id)

                if ionomycin_response:
                    peak = self.dff_data.loc[well_id].max()
                    normalized = (peak / ionomycin_response) * 100

                    ws.cell(row=row, column=1, value=group_name)
                    ws.cell(row=row, column=2, value=well_id)
                    ws.cell(row=row, column=3, value=concentration)
                    ws.cell(row=row, column=4, value=float(normalized))
                    ws.cell(row=row, column=5, value=sample_id)
                    ws.cell(row=row, column=6, value=float(ionomycin_response))
                    row += 1

    def create_analysis_metrics_sheet(self, wb):
        """Create new sheet with detailed analysis metrics"""
        ws = wb.create_sheet("Analysis_Metrics")

        # Add headers
        headers = [
            "Group", "Metric", "Mean", "SEM",
            "Min", "Max", "N"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        # Add data
        row = 2
        grouped_data = self.group_data_by_metadata()

        for group_name, well_ids in grouped_data.items():
            group_data = self.dff_data.loc[well_ids]

            # Calculate metrics
            peaks = group_data.max(axis=1)
            peak_times = group_data.idxmax(axis=1).astype(float)
            group_auc = self.auc_data[well_ids]

            # Add peak response metrics
            metrics = [
                ("Peak ΔF/F₀", peaks),
                ("Time to Peak (s)", peak_times),
                ("AUC", group_auc)
            ]

            for metric_name, values in metrics:
                ws.cell(row=row, column=1, value=group_name)
                ws.cell(row=row, column=2, value=metric_name)
                ws.cell(row=row, column=3, value=float(values.mean()))
                ws.cell(row=row, column=4, value=float(values.std() / np.sqrt(len(values))))
                ws.cell(row=row, column=5, value=float(values.min()))
                ws.cell(row=row, column=6, value=float(values.max()))
                ws.cell(row=row, column=7, value=len(values))
                row += 1

            # Add blank row between groups
            row += 1

    def setup_plot_controls(self, layout):
        """Set up plot control buttons"""
        plot_buttons = QHBoxLayout()

        self.raw_plot_button = QPushButton("Raw Traces")
        self.dff_plot_button = QPushButton("ΔF/F₀ Traces")
        self.summary_plot_button = QPushButton("Summary Plots")

        self.raw_plot_button.setCheckable(True)
        self.dff_plot_button.setCheckable(True)
        self.summary_plot_button.setCheckable(True)

        self.raw_plot_button.clicked.connect(lambda: self.toggle_plot_window('raw'))
        self.dff_plot_button.clicked.connect(lambda: self.toggle_plot_window('dff'))
        self.summary_plot_button.clicked.connect(lambda: self.toggle_plot_window('summary'))

        plot_buttons.addWidget(self.raw_plot_button)
        plot_buttons.addWidget(self.dff_plot_button)
        plot_buttons.addWidget(self.summary_plot_button)

        layout.addLayout(plot_buttons)

    def create_menus(self):
        """Create menu bar and menus"""
        menubar = self.menuBar()

        # Add File menu
        file_menu = menubar.addMenu('File')

        # Add export FLIPR layout action
        export_flipr_action = QAction('Export FLIPR Layout...', self)
        export_flipr_action.triggered.connect(self.export_to_flipr)
        file_menu.addAction(export_flipr_action)

        # Add separator
        file_menu.addSeparator()

        # Analysis menu
        analysis_menu = menubar.addMenu('Analysis')

        # Parameters action
        params_action = QAction('Parameters...', self)
        params_action.triggered.connect(self.show_parameters_dialog)
        analysis_menu.addAction(params_action)

        # Help menu
        help_menu = menubar.addMenu('Help')

        # About action
        about_action = QAction('About', self)
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)

        # Manual action
        manual_action = QAction('User Manual', self)
        manual_action.triggered.connect(self.show_manual)
        help_menu.addAction(manual_action)

    def show_parameters_dialog(self):
        """Show parameters dialog and update values if accepted"""
        dialog = ParametersDialog(self)

        # Set current values
        dialog.artifact_start.setValue(self.analysis_params['artifact_start'])
        dialog.artifact_end.setValue(self.analysis_params['artifact_end'])
        dialog.baseline_frames.setValue(self.analysis_params['baseline_frames'])
        dialog.peak_start_frame.setValue(self.analysis_params['peak_start_frame'])

        if dialog.exec_():
            # Update parameters if dialog is accepted
            self.analysis_params['artifact_start'] = dialog.artifact_start.value()
            self.analysis_params['artifact_end'] = dialog.artifact_end.value()
            self.analysis_params['baseline_frames'] = dialog.baseline_frames.value()
            self.analysis_params['peak_start_frame'] = dialog.peak_start_frame.value()

            # Reprocess data if needed
            if self.raw_data is not None:
                self.process_data()
                self.update_plots()

    def show_about(self):
        """Show about dialog"""
        QMessageBox.about(self, "About FLIPR Analysis",
            "FLIPR Analysis Tool\n\n"
            "A tool for analyzing calcium imaging data from FLIPR experiments.\n\n"
            "Version 1.0")

    def show_manual(self):
        """Show user manual"""
        QMessageBox.information(self, "FLIPR Analysis Tool - Quick Guide",
        """FLIPR Analysis Tool
        Quick Start Guide:

        Loading Data
        • Click 'Load Data' to import your .seq1 file
        • The well plate grid will become active
        Well Selection
        • Click individual wells to select/deselect
        • Click row (A-H) or column (1-12) headers to select entire rows/columns
        • Click top-left corner to select all wells
        Labeling Wells
        • Enter agonist name, concentration, and sample ID
        • Use color selector to assign group colors
        • Select wells and click 'Apply Label'
        • For dilution series, use 'Log10 Series' mode
        Analysis Options
        • Access parameters via Analysis → Parameters menu
        • Toggle 'Remove Injection Artifact' to clean data
        • Enable 'Normalize to Ionomycin' for normalized responses
        Visualization (Summary Plots)
        • Individual Traces: All ΔF/F₀ traces by group
        • Mean Traces: Average responses with SEM
        • Peak Responses: Maximum responses with error bars
        • Normalized Responses: Data as % of ionomycin (when enabled)

        Tips:

        Label wells before analysis for proper grouping
        Set analysis parameters before processing
        Use consistent naming for proper grouping
        Save layouts for repeated analysis

        For more detailed information, please refer to the full user manual.""")


    def show_plot_window(self):
        """Show the plot window and plot selected wells"""
        if not self.raw_data is not None:
            QMessageBox.warning(self, "Warning", "Please load data first")
            return

        self.plot_window.show()
        self.update_plots()

    def toggle_artifact_removal(self, state):
        """Toggle artifact removal and update plots"""
        self.remove_artifact = bool(state)
        if self.raw_data is not None:
            self.process_data()
            self.update_plots()

    def toggle_ionomycin_normalization(self, state):
        """Toggle ionomycin normalization and update plots"""
        self.normalize_to_ionomycin = bool(state)
        if self.raw_data is not None:
            self.update_summary_plots()

    def toggle_plot_window(self, plot_type):
        """Toggle visibility of specified plot window"""
        if self.raw_data is None:
            QMessageBox.warning(self, "Warning", "Please load data first")
            return

        windows = {
            'raw': (self.raw_plot_window, self.raw_plot_button),
            'dff': (self.dff_plot_window, self.dff_plot_button),
            'summary': (self.summary_plot_window, self.summary_plot_button)
        }

        window, button = windows[plot_type]

        if window.isVisible():
            window.hide()
            button.setChecked(False)
        else:
            # Process data if needed
            if plot_type in ['dff', 'summary'] and self.dff_data is None:
                self.process_data()

            window.show()
            button.setChecked(True)

            # Update appropriate plots
            if plot_type == 'summary':
                logger.info("Triggering summary plot update...")
                self.update_summary_plots()
            else:
                self.update_plots()

    def group_data_by_metadata(self):
        """Group data based on available metadata"""
        grouped_data = {}

        # Default group for all wells if no metadata
        all_wells = []

        for idx in range(96):
            well_data = self.well_data[idx]
            well_id = well_data["well_id"]
            all_wells.append(well_id)

            # Create grouping key based on available metadata
            key_parts = []
            if well_data.get("label"):  # Agonist
                key_parts.append(well_data["label"])
            if well_data.get("concentration"):
                key_parts.append(f"{well_data['concentration']}")
            if well_data.get("sample_id"):
                key_parts.append(well_data["sample_id"])

            if key_parts:  # If we have metadata, use it for grouping
                group_key = " | ".join(key_parts)
                if group_key not in grouped_data:
                    grouped_data[group_key] = []
                grouped_data[group_key].append(well_id)

        # If no groups were created, use all wells as a single group
        if not grouped_data:
            grouped_data["All Wells"] = all_wells

        logger.info(f"Created {len(grouped_data)} groups: {list(grouped_data.keys())}")
        return grouped_data

    def get_ionomycin_responses(self):
        """Calculate mean ionomycin responses for each sample ID"""
        ionomycin_responses = {}

        # Group wells by sample ID
        sample_groups = {}
        for idx in range(96):
            well_data = self.well_data[idx]
            if well_data["label"] == "Ionomycin":
                sample_id = well_data.get("sample_id", "default")
                if sample_id not in sample_groups:
                    sample_groups[sample_id] = []
                sample_groups[sample_id].append(well_data["well_id"])

        # Calculate mean ionomycin response for each sample
        for sample_id, wells in sample_groups.items():
            peaks = self.dff_data.loc[wells].max(axis=1)
            ionomycin_responses[sample_id] = peaks.mean()

        return ionomycin_responses


    def update_summary_plots(self):
        """Update summary plots based on grouped data"""
        self.show_status("Updating plots...")
        logger.info("Starting summary plot update...")

        if self.dff_data is None:
            logger.info("Processing data for summary plots...")
            self.process_data()

        # Clear existing plots
        self.summary_plot_window.clear_plots()

        # Get appropriate time values that match the processed data length
        if self.remove_artifact:
            times = self.processed_time_points  # Use already processed time points
        else:
            times = np.array(pd.to_numeric(self.raw_data.columns, errors='coerce'))

        logger.info(f"Time points shape: {times.shape}")
        logger.info(f"Data shape: {self.dff_data.shape}")

        # Get grouped data
        grouped_data = self.group_data_by_metadata()
        logger.info(f"Processing {len(grouped_data)} groups for plotting")

        # Add legend to individual traces plot
        self.summary_plot_window.individual_plot.addLegend(offset=(10, 10))

        # Add legend to mean traces plot
        self.summary_plot_window.mean_plot.addLegend(offset=(10, 10))

        # Plot traces for each group
        non_ionomycin_count = 0  # Counter for normalized plot positioning
        for i, (group_name, well_ids) in enumerate(grouped_data.items()):
            logger.info(f"Plotting group '{group_name}' with {len(well_ids)} wells")

            try:
                # Get color for this group
                well_idx = next(idx for idx in range(96) if self.well_data[idx]["well_id"] == well_ids[0])
                base_color = self.well_data[well_idx]["color"]
                main_color = QColor(base_color)
                transparent_color = QColor(base_color)
                transparent_color.setAlpha(50)
                light_pen = pg.mkPen(color=main_color, width=1, alpha=50)

                # Get group data
                group_data = self.dff_data.loc[well_ids]

                # Plot individual traces with proper legend
                for well_id in well_ids:
                    trace_data = np.array(self.dff_data.loc[well_id])
                    if len(times) == len(trace_data):
                        self.summary_plot_window.individual_plot.plot(
                            times,
                            trace_data,
                            pen=light_pen,
                            name=group_name if well_id == well_ids[0] else None  # Legend only for first trace
                        )

                # Calculate and plot mean trace
                mean_trace = np.array(group_data.mean())
                sem_trace = np.array(group_data.sem())

                if len(times) == len(mean_trace):
                    # Plot mean trace
                    self.summary_plot_window.mean_plot.plot(
                        times,
                        mean_trace,
                        pen=pg.mkPen(main_color, width=2),
                        name=group_name
                    )

                    # Add error bands
                    band_top = mean_trace + sem_trace
                    band_bottom = mean_trace - sem_trace
                    fill = pg.FillBetweenItem(
                        pg.PlotDataItem(times, band_top),
                        pg.PlotDataItem(times, band_bottom),
                        brush=transparent_color
                    )
                    self.summary_plot_window.mean_plot.addItem(fill)

                # Calculate peak responses
                peaks = np.array(group_data.max(axis=1))
                peak_mean = np.mean(peaks)
                peak_sem = np.std(peaks) / np.sqrt(len(peaks))

                # Add bar for peak response
                bar = pg.BarGraphItem(
                    x=[i],
                    height=[peak_mean],
                    width=0.8,
                    brush=main_color
                )
                self.summary_plot_window.responses_plot.addItem(bar)

                # Add error bars
                error = pg.ErrorBarItem(
                    x=np.array([i]),
                    y=np.array([peak_mean]),
                    height=np.array([peak_sem * 2]),
                    beam=0.2,
                    pen=pg.mkPen(main_color, width=2)
                )
                self.summary_plot_window.responses_plot.addItem(error)

                # Add value label above bar
                text = pg.TextItem(
                    text=f'{peak_mean:.1f}±{peak_sem:.1f}',
                    color=main_color,
                    anchor=(0.5, 1)
                )
                text.setPos(i, peak_mean + peak_sem * 2)
                self.summary_plot_window.responses_plot.addItem(text)

                # Add normalized responses if enabled (only for non-ionomycin groups)
                if self.normalize_to_ionomycin and "ionomycin" not in group_name.lower():
                    ionomycin_responses = self.get_ionomycin_responses()
                    if ionomycin_responses:
                        normalized_peaks = []
                        for well_id in well_ids:
                            well_idx = next(idx for idx in range(96) if self.well_data[idx]["well_id"] == well_id)
                            sample_id = self.well_data[well_idx].get("sample_id", "default")
                            ionomycin_response = ionomycin_responses.get(sample_id)
                            if ionomycin_response:
                                peak = group_data.loc[well_id].max()
                                normalized_peaks.append((peak / ionomycin_response) * 100)

                        if normalized_peaks:
                            norm_mean = np.mean(normalized_peaks)
                            norm_sem = np.std(normalized_peaks) / np.sqrt(len(normalized_peaks))

                            # Add normalized bar with matching color
                            norm_bar = pg.BarGraphItem(
                                x=[non_ionomycin_count],
                                height=[norm_mean],
                                width=0.8,
                                brush=main_color
                            )
                            self.summary_plot_window.normalized_plot.addItem(norm_bar)

                            # Add normalized error bars
                            norm_error = pg.ErrorBarItem(
                                x=np.array([non_ionomycin_count]),
                                y=np.array([norm_mean]),
                                height=np.array([norm_sem * 2]),
                                beam=0.2,
                                pen=pg.mkPen(main_color, width=2)
                            )
                            self.summary_plot_window.normalized_plot.addItem(norm_error)

                            # Add normalized value label
                            norm_text = pg.TextItem(
                                text=f'{norm_mean:.1f}±{norm_sem:.1f}',
                                color=main_color,
                                anchor=(0.5, 1)
                            )
                            norm_text.setPos(non_ionomycin_count, norm_mean + norm_sem * 2)
                            self.summary_plot_window.normalized_plot.addItem(norm_text)

                            non_ionomycin_count += 1  # Increment counter for next non-ionomycin group

                # Calculate AUC for this group
                group_auc = self.auc_data[well_ids]
                auc_mean = group_auc.mean()
                auc_sem = group_auc.std() / np.sqrt(len(group_auc))

                # Calculate time to peak
                peak_times = group_data.idxmax(axis=1).astype(float)
                time_to_peak_mean = peak_times.mean()
                time_to_peak_sem = peak_times.std() / np.sqrt(len(peak_times))

                # Add AUC bar
                auc_bar = pg.BarGraphItem(
                    x=[i],
                    height=[auc_mean],
                    width=0.8,
                    brush=main_color
                )
                self.summary_plot_window.auc_plot.addItem(auc_bar)

                # Add AUC error bars
                auc_error = pg.ErrorBarItem(
                    x=np.array([i]),
                    y=np.array([auc_mean]),
                    height=np.array([auc_sem * 2]),
                    beam=0.2,
                    pen=pg.mkPen(main_color, width=2)
                )
                self.summary_plot_window.auc_plot.addItem(auc_error)

                # Add AUC value label
                auc_text = pg.TextItem(
                    text=f'{auc_mean:.1f}±{auc_sem:.1f}',
                    color=main_color,
                    anchor=(0.5, 1)
                )
                auc_text.setPos(i, auc_mean + auc_sem * 2)
                self.summary_plot_window.auc_plot.addItem(auc_text)

                # Add Time to Peak bar
                ttp_bar = pg.BarGraphItem(
                    x=[i],
                    height=[time_to_peak_mean],
                    width=0.8,
                    brush=main_color
                )
                self.summary_plot_window.time_to_peak_plot.addItem(ttp_bar)

                # Add Time to Peak error bars
                ttp_error = pg.ErrorBarItem(
                    x=np.array([i]),
                    y=np.array([time_to_peak_mean]),
                    height=np.array([time_to_peak_sem * 2]),
                    beam=0.2,
                    pen=pg.mkPen(main_color, width=2)
                )
                self.summary_plot_window.time_to_peak_plot.addItem(ttp_error)

                # Add Time to Peak value label
                ttp_text = pg.TextItem(
                    text=f'{time_to_peak_mean:.1f}±{time_to_peak_sem:.1f}',
                    color=main_color,
                    anchor=(0.5, 1)
                )
                ttp_text.setPos(i, time_to_peak_mean + time_to_peak_sem * 2)
                self.summary_plot_window.time_to_peak_plot.addItem(ttp_text)



                logger.info(f"Successfully plotted group {group_name}")


            except Exception as e:
                logger.error(f"Error plotting group {group_name}: {str(e)}")
                import traceback
                logger.error(traceback.format_exc())
                continue

        # Update all plot labels and settings
        for plot in [self.summary_plot_window.individual_plot,
                    self.summary_plot_window.mean_plot]:
            plot.setLabel('left', 'ΔF/F₀')
            plot.setLabel('bottom', 'Time (s)')

        # Update peak responses plot
        self.summary_plot_window.responses_plot.setLabel('left', 'Peak ΔF/F₀')
        self.summary_plot_window.responses_plot.setLabel('bottom', 'Groups')

        # Add group labels to response plot (all groups)
        group_names = list(grouped_data.keys())
        axis = self.summary_plot_window.responses_plot.getAxis('bottom')
        ticks = [(i, name) for i, name in enumerate(group_names)]
        axis.setTicks([ticks])

        # Set axis ranges for responses plot
        n_groups = len(grouped_data)
        self.summary_plot_window.responses_plot.setXRange(-0.5, n_groups - 0.5)

        # Update normalized plot labels and axes
        if self.normalize_to_ionomycin:
            self.summary_plot_window.normalized_plot.setLabel('left', 'Response (% Ionomycin)')
            self.summary_plot_window.normalized_plot.setLabel('bottom', 'Groups')

            # Create ticks only for non-ionomycin groups
            non_ionomycin_groups = [name for name in grouped_data.keys()
                                  if "ionomycin" not in name.lower()]
            norm_ticks = [(i, name) for i, name in enumerate(non_ionomycin_groups)]

            axis = self.summary_plot_window.normalized_plot.getAxis('bottom')
            axis.setTicks([norm_ticks])
            self.summary_plot_window.normalized_plot.setXRange(-0.5, len(non_ionomycin_groups) - 0.5)

        # Update AUC plot labels and axes
        axis = self.summary_plot_window.auc_plot.getAxis('bottom')
        ticks = [(i, name) for i, name in enumerate(grouped_data.keys())]
        axis.setTicks([ticks])
        self.summary_plot_window.auc_plot.setXRange(-0.5, len(grouped_data) - 0.5)

        # Update Time to Peak plot labels and axes
        axis = self.summary_plot_window.time_to_peak_plot.getAxis('bottom')
        axis.setTicks([ticks])
        self.summary_plot_window.time_to_peak_plot.setXRange(-0.5, len(grouped_data) - 0.5)


        # Update results text
        self.update_results_text()

        self.show_status("Plots updated", 3000)
        logger.info("Summary plot update completed")


    def open_file_dialog(self):
        """Open file dialog to load FLIPR data"""
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(self, "Open Data File", "",
                                                 "Text Files (*.txt *.seq1);;All Files (*)", options=options)
        if file_path:
            try:
                self.show_status("Loading data...")
                self.raw_data, self.original_filename = self.load_data(file_path)
                self.file_display.setText(self.original_filename)
                self.show_status(f"Data loaded successfully", 3000)
            except Exception as e:
                self.show_status(f"Error: {str(e)}", 5000)
                QMessageBox.critical(self, "Error", f"Failed to load data: {str(e)}")



    def load_data(self, file_path: str) -> Tuple[pd.DataFrame, str]:
        """Load and preprocess FLIPR data from file."""
        try:
            # Validate file extension
            if not file_path.endswith(('.txt', '.seq1')):
                raise ValueError("Invalid file format. Must be .txt or .seq1")

            # Read header
            with open(file_path, 'r') as f:
                header = f.readline().strip().split('\t')
            original_filename = header[0]
            header_values = header[5::]  # Time points start from column 5

            # Read data
            with open(file_path, 'r') as file:
                reader = csv.reader(file, delimiter='\t')
                next(reader)  # Skip header
                raw_data = []
                for row in reader:
                    if len(row) > 4:  # Ensure we have enough columns
                        raw_data.append(row[4:])  # Skip first 4 columns

            # Create DataFrame
            data = pd.DataFrame(raw_data)
            data.set_index(0, inplace=True)
            data.index.name = 'Well'

            # Remove any extra columns
            if len(data.columns) > len(header_values):
                data = data.iloc[:, :len(header_values)]

            data.columns = header_values

            # Convert to numeric, replacing any non-numeric values with NaN
            data = data.apply(pd.to_numeric, errors='coerce')

            # Reset processed data
            self.dff_data = None
            self.zeroed_data = None

            # Debug logging
            logger.info(f"Loaded data shape: {data.shape}")
            logger.info(f"Sample of loaded data:\n{data.iloc[:3, :5]}")  # Show first 3 rows, 5 columns

            return data, original_filename

        except Exception as e:
            logger.error(f"Error loading data from {file_path}: {str(e)}")
            raise


    def plot_single_well(self, idx):
        """Plot a single well's trace"""
        well_id = self.well_data[idx]["well_id"]

        if well_id in self.raw_data.index:
            try:
                times = pd.to_numeric(self.raw_data.columns, errors='coerce')
                values = pd.to_numeric(self.raw_data.loc[well_id], errors='coerce')

                if not values.isna().all():
                    color = self.well_data[idx]["color"]
                    self.plot_window.plot_trace(well_id, times, values, color)
            except Exception as e:
                logger.error(f"Error plotting well {well_id}: {str(e)}")

    def update_traces_for_selection_change(self, previously_selected):
        """Helper function to update plot traces when selection changes"""
        try:
            if self.raw_data is None:
                logger.warning("Attempted to update traces with no data loaded")
                return

            newly_selected = self.selected_wells - previously_selected
            newly_unselected = previously_selected - self.selected_wells

            logger.info(f"Updating traces: {len(newly_selected)} wells to add, {len(newly_unselected)} wells to remove")

            # Remove traces for unselected wells
            self.remove_traces(newly_unselected)

            # Add traces for newly selected wells
            self.add_traces(newly_selected)

            # Update summary plots if they exist
            if hasattr(self, 'summary_plot_window'):
                self.update_summary_plots()

        except Exception as e:
            logger.error(f"Error updating traces: {str(e)}")
            logger.debug(f"Stack trace:", exc_info=True)
            QMessageBox.warning(self, "Error", "Failed to update plot traces. See log for details.")

    def remove_traces(self, indices):
        """Remove traces for given well indices"""
        for idx in indices:
            try:
                well_id = self.well_data[idx]["well_id"]
                logger.debug(f"Removing traces for well {well_id}")

                # Remove from raw plot
                if hasattr(self, 'raw_plot_window'):
                    if well_id in self.raw_plot_window.plot_items:
                        self.raw_plot_window.plot_widget.removeItem(self.raw_plot_window.plot_items[well_id])
                        del self.raw_plot_window.plot_items[well_id]

                # Remove from ΔF/F₀ plot
                if hasattr(self, 'dff_plot_window'):
                    if well_id in self.dff_plot_window.plot_items:
                        self.dff_plot_window.plot_widget.removeItem(self.dff_plot_window.plot_items[well_id])
                        del self.dff_plot_window.plot_items[well_id]

            except KeyError as e:
                logger.error(f"Well data not found for index {idx}: {str(e)}")
            except Exception as e:
                logger.error(f"Error removing traces for well index {idx}: {str(e)}")
                logger.debug("Stack trace:", exc_info=True)

    def add_traces(self, indices):
        """Add traces for given well indices"""
        try:
            times = self.get_time_points()
        except Exception as e:
            logger.error(f"Failed to get time points: {str(e)}")
            logger.debug("Stack trace:", exc_info=True)
            return

        for idx in indices:
            try:
                well_id = self.well_data[idx]["well_id"]
                if well_id not in self.raw_data.index:
                    logger.warning(f"Well {well_id} not found in raw data")
                    continue

                logger.debug(f"Adding traces for well {well_id}")

                # Add to raw plot
                if hasattr(self, 'raw_plot_window'):
                    try:
                        values = self.get_raw_values(well_id)
                        self.raw_plot_window.plot_trace(well_id, times, values, self.well_data[idx]["color"])
                    except Exception as e:
                        logger.error(f"Error plotting raw trace for well {well_id}: {str(e)}")

                # Add to ΔF/F₀ plot
                if hasattr(self, 'dff_plot_window'):
                    try:
                        if self.dff_data is None:
                            logger.info("Processing data for ΔF/F₀ calculation")
                            self.process_data()
                        if well_id in self.dff_data.index:
                            values = self.dff_data.loc[well_id]
                            self.dff_plot_window.plot_trace(well_id, times, values, self.well_data[idx]["color"])
                        else:
                            logger.warning(f"Well {well_id} not found in ΔF/F₀ data")
                    except Exception as e:
                        logger.error(f"Error plotting ΔF/F₀ trace for well {well_id}: {str(e)}")

            except Exception as e:
                logger.error(f"Error adding traces for well index {idx}: {str(e)}")
                logger.debug("Stack trace:", exc_info=True)

    def get_time_points(self):
        """Get appropriate time points based on artifact removal setting"""
        try:
            if self.remove_artifact:
                if self.processed_time_points is None:
                    raise ValueError("Processed time points not available")
                return self.processed_time_points

            times = pd.to_numeric(self.raw_data.columns, errors='coerce')
            if times.isna().any():
                logger.warning("Some time points could not be converted to numeric values")
            return times

        except Exception as e:
            logger.error(f"Error getting time points: {str(e)}")
            logger.debug("Stack trace:", exc_info=True)
            raise

    def get_raw_values(self, well_id):
        """Get raw values for a well, handling artifact removal if enabled"""
        try:
            if self.remove_artifact:
                n_cols = self.raw_data.shape[1]
                start_idx = int(n_cols * self.analysis_params['artifact_start']/220)
                end_idx = int(n_cols * self.analysis_params['artifact_end']/220)

                if start_idx >= end_idx:
                    raise ValueError("Invalid artifact removal indices")

                values = pd.concat([
                    self.raw_data.loc[well_id][:start_idx],
                    self.raw_data.loc[well_id][end_idx:]
                ])
                logger.debug(f"Artifact removed for well {well_id}: frames {start_idx}-{end_idx}")
                return values

            return self.raw_data.loc[well_id]

        except KeyError:
            logger.error(f"Well {well_id} not found in data")
            raise
        except Exception as e:
            logger.error(f"Error getting raw values for well {well_id}: {str(e)}")
            logger.debug("Stack trace:", exc_info=True)
            raise

    def get_row_col(self, index):
        """Convert well index to row and column"""
        row = index // 12  # 12 columns per row
        col = index % 12
        return row, col

    def get_index(self, row, col):
        """Convert row and column to well index"""
        if 0 <= row < 8 and 0 <= col < 12:  # Check bounds
            return row * 12 + col
        return None

    def select_rectangle(self, start_index, end_index):
        """Select all wells within the rectangle defined by start and end indices"""
        # Get corners of the rectangle
        start_row, start_col = self.get_row_col(start_index)
        end_row, end_col = self.get_row_col(end_index)

        # Determine rectangle bounds
        min_row = min(start_row, end_row)
        max_row = max(start_row, end_row)
        min_col = min(start_col, end_col)
        max_col = max(start_col, end_col)

        print(f"Selecting rectangle from ({min_row}, {min_col}) to ({max_row}, {max_col})")

        # Clear previous selection if not in wells mode
        if not self.selection_state['wells']:
            self.selection_state['wells'] = set()

        # Select all wells within the rectangle
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                idx = self.get_index(row, col)
                if idx is not None:
                    self.selection_state['wells'].add(idx)

        # Update display
        self.update_selection_state()
        self.update_well_appearances()

    def toggle_well_selection(self, index):
        """Handle individual well selection"""
        modifiers = QApplication.keyboardModifiers()
        previously_selected = set(self.selected_wells)

        if modifiers == Qt.ShiftModifier:
            # For shift selection, add rectangle to existing selection
            if self.last_selected is not None:
                self.add_rectangle_to_selection(self.last_selected, index)
            self.last_selected = index
        else:
            # Toggle individual well
            if index in self.selection_state['wells']:
                self.selection_state['wells'].remove(index)
            else:
                # Convert any row/column selections to individual wells first
                if self.selection_state['rows'] or self.selection_state['cols']:
                    self.convert_to_well_selection()
                self.selection_state['wells'].add(index)
            self.last_selected = index

        self.update_selection_state()
        self.update_well_appearances()
        self.update_traces_for_selection_change(previously_selected)

    def convert_to_well_selection(self):
        """Convert row/column selections to individual well selections"""
        # Add wells from rows
        for row in self.selection_state['rows']:
            self.selection_state['wells'].update(range(row * 12, (row + 1) * 12))

        # Add wells from columns
        for col in self.selection_state['cols']:
            self.selection_state['wells'].update(range(col, 96, 12))

        # Clear row and column selections
        self.selection_state['rows'] = set()
        self.selection_state['cols'] = set()

    def add_rectangle_to_selection(self, start_index, end_index):
        """Add wells within rectangle to selection without clearing existing selection"""
        # Convert any row/column selections to individual wells first
        self.convert_to_well_selection()

        # Get corners of the rectangle
        start_row, start_col = self.get_row_col(start_index)
        end_row, end_col = self.get_row_col(end_index)

        # Determine rectangle bounds
        min_row = min(start_row, end_row)
        max_row = max(start_row, end_row)
        min_col = min(start_col, end_col)
        max_col = max(start_col, end_col)

        print(f"Adding rectangle from ({min_row}, {min_col}) to ({max_row}, {max_col})")

        # Add all wells within the rectangle to selection
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                idx = self.get_index(row, col)
                if idx is not None:
                    self.selection_state['wells'].add(idx)

    def toggle_row_selection(self, row_index):
        """Toggle row selection"""
        previously_selected = set(self.selected_wells)

        # Convert existing row/column selections to individual wells
        self.convert_to_well_selection()

        # Toggle all wells in this row
        row_wells = set(range(row_index * 12, (row_index + 1) * 12))
        if row_wells.issubset(self.selection_state['wells']):
            # If all wells in row are selected, remove them
            self.selection_state['wells'] -= row_wells
        else:
            # Otherwise add them
            self.selection_state['wells'].update(row_wells)

        self.update_selection_state()
        self.update_well_appearances()
        self.update_traces_for_selection_change(previously_selected)

    def toggle_column_selection(self, col_index):
        """Toggle column selection"""
        previously_selected = set(self.selected_wells)

        # Convert existing row/column selections to individual wells
        self.convert_to_well_selection()

        # Toggle all wells in this column
        col_wells = set(range(col_index, 96, 12))
        if col_wells.issubset(self.selection_state['wells']):
            # If all wells in column are selected, remove them
            self.selection_state['wells'] -= col_wells
        else:
            # Otherwise add them
            self.selection_state['wells'].update(col_wells)

        self.update_selection_state()
        self.update_well_appearances()
        self.update_traces_for_selection_change(previously_selected)

    def toggle_all_selection(self):
        """Toggle selection of all wells"""
        previously_selected = set(self.selected_wells)

        if len(self.selected_wells) == 96:
            # If all wells are selected, clear selection
            self.selection_state = {
                'rows': set(),
                'cols': set(),
                'wells': set(),
                'all_selected': False
            }
        else:
            # Otherwise select all wells
            self.selection_state = {
                'rows': set(),
                'cols': set(),
                'wells': set(range(96)),
                'all_selected': False
            }

        self.update_selection_state()
        self.update_well_appearances()
        self.update_traces_for_selection_change(previously_selected)
    def update_selection_state(self):
        """Update the selected_wells set based on current selection state"""
        try:
            # Create a new set for selected wells
            selected = set()

            # If all wells are selected, that takes precedence
            if self.selection_state['all_selected']:
                selected = set(range(96))
            else:
                # Add wells from rows
                for row in self.selection_state['rows']:
                    selected.update(range(row * 12, (row + 1) * 12))

                # Add wells from columns
                for col in self.selection_state['cols']:
                    selected.update(range(col, 96, 12))

                # Add individual wells
                selected.update(self.selection_state['wells'])

            # Update the selected_wells set
            self.selected_wells = selected

        except Exception as e:
            logger.error(f"Error updating selection state: {str(e)}")
            QMessageBox.warning(self, "Error",
                              "Failed to update well selection. See log for details.")


    def apply_label(self):
        """Apply label, concentration, and color to selected wells"""
        mode = self.mode_selector.currentText()

        if mode == "Simple Label":
            for idx in self.selected_wells:
                # Only update fields that are checked
                if self.label_checkbox.isChecked():
                    self.well_data[idx]["label"] = self.label_input.text()

                if self.concentration_checkbox.isChecked():
                    concentration = self.starting_conc_input.text()
                    self.well_data[idx]["concentration"] = f"{concentration} µM" if concentration else ""

                if self.sample_id_checkbox.isChecked():
                    self.well_data[idx]["sample_id"] = self.sample_id_input.text()

                if self.color_checkbox.isChecked():
                    if self.current_color.name() != self.default_colors[idx % len(self.default_colors)]:
                        self.well_data[idx]["color"] = self.current_color.name()

                self.update_button(idx)

        elif mode == "Log10 Series":
            try:
                if not self.concentration_checkbox.isChecked():
                    QMessageBox.warning(self, "Input Error", "Concentration must be enabled for Log10 Series")
                    return

                starting_conc = float(self.starting_conc_input.text())
            except ValueError:
                QMessageBox.warning(self, "Input Error", "Invalid starting concentration")
                return

            if len(self.selected_wells) < 2:
                QMessageBox.warning(self, "Selection Error", "Select at least 2 wells for Log10 Series")
                return

            # Calculate log10 series concentrations
            sorted_indices = sorted(self.selected_wells)
            num_wells = len(sorted_indices)
            concentrations = [starting_conc / (10 ** i) for i in range(num_wells)]

            for idx, conc in zip(sorted_indices, concentrations):
                if self.concentration_checkbox.isChecked():
                    self.well_data[idx]["concentration"] = f"{conc:.2f} µM"
                if self.label_checkbox.isChecked():
                    self.well_data[idx]["label"] = self.label_input.text()
                if self.sample_id_checkbox.isChecked():
                    self.well_data[idx]["sample_id"] = self.sample_id_input.text()
                if self.color_checkbox.isChecked():
                    if self.current_color.name() != self.default_colors[idx % len(self.default_colors)]:
                        self.well_data[idx]["color"] = self.current_color.name()
                self.update_button(idx)

        elif mode == "Clear Wells":
            for idx in self.selected_wells:
                # Only clear fields that are checked
                if self.label_checkbox.isChecked():
                    self.well_data[idx]["label"] = ""
                if self.concentration_checkbox.isChecked():
                    self.well_data[idx]["concentration"] = ""
                if self.sample_id_checkbox.isChecked():
                    self.well_data[idx]["sample_id"] = ""
                if self.color_checkbox.isChecked():
                    default_color = self.default_colors[idx % len(self.default_colors)]
                    self.well_data[idx]["color"] = default_color
                self.update_button(idx)

        # Update all visible plot windows
        if self.raw_plot_window.isVisible():
            self.update_plots()
        if self.dff_plot_window.isVisible():
            if self.dff_data is None:
                self.process_data()
            self.update_plots()
        if self.summary_plot_window.isVisible():
            self.update_summary_plots()

        self.selected_wells.clear()

    def update_button(self, idx):
        """Update button text and color"""
        data = self.well_data[idx]
        text = f"{data['well_id']}"

        if data['label']:
            text += f"\n{data['label']}"
        if data['concentration']:
            text += f"\n{data['concentration']}"
        if data['sample_id']:
            text += f"\n{data['sample_id']}"

        self.wells[idx].setText(text)
        self.wells[idx].setStyleSheet(
            f"background-color: {data['color']}; color: black;"  # Ensure text is always black
        )

    def clear_selection(self):
        """Clear the selection and reset buttons to default state"""
        for idx in self.selected_wells:
            # Get the default color for this index
            default_color = self.default_colors[idx % len(self.default_colors)]

            # Reset the well data but preserve the well_id
            well_id = self.well_data[idx]["well_id"]
            self.well_data[idx] = {
                "well_id": well_id,
                "label": "",
                "concentration": "",
                "sample_id": "",
                "color": default_color
            }

            # Update button appearance
            self.wells[idx].setText(well_id)  # Reset text to just the well ID
            self.wells[idx].setStyleSheet(f"background-color: {default_color}; color: black;")

        # Clear the selection set
        self.selected_wells.clear()

        # Update plots if plot window is visible
        if self.plot_window.isVisible():
            self.update_plots()

    def save_layout(self):
        """Save the current layout to a JSON file"""
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(self, "Save Layout", "", "JSON Files (*.json)", options=options)
        if file_path:
            with open(file_path, "w") as f:
                json.dump(self.well_data, f)

    def load_layout(self):
        """Load a layout from a JSON file"""
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(self, "Load Layout", "", "JSON Files (*.json)", options=options)
        if file_path:
            with open(file_path, "r") as f:
                self.well_data = json.load(f)
            for idx, data in enumerate(self.well_data):
                self.update_button(idx)

    def process_data(self):
        """Process loaded data with artifact removal if enabled"""
        if self.raw_data is None:
            return

        try:
            self.show_status("Processing data...")

            # Work with a copy of the raw data
            processed_data = self.raw_data.copy()

            # Initialize time points
            all_time_points = pd.to_numeric(processed_data.columns, errors='coerce')

            # Remove artifact if enabled
            if self.remove_artifact:
                n_cols = processed_data.shape[1]
                start_idx = int(n_cols * self.analysis_params['artifact_start']/220)
                end_idx = int(n_cols * self.analysis_params['artifact_end']/220)

                # Store the time points before removing data
                self.processed_time_points = np.concatenate([
                    all_time_points[:start_idx],
                    all_time_points[end_idx:]
                ])

                # Remove artifact points from data
                processed_data = processed_data.drop(processed_data.columns[start_idx:end_idx], axis=1)
                processed_data.columns = self.processed_time_points  # Update column names
            else:
                # If no artifact removal, use all time points
                self.processed_time_points = all_time_points

            # Calculate F0
            F0 = self.processor.get_F0(processed_data,
                                     baseline_frames=self.analysis_params['baseline_frames'])

            # Calculate ΔF/F₀
            self.dff_data = self.processor.calculate_dff(processed_data, F0)

            # Calculate AUC for ΔF/F₀ traces
            self.auc_data = self.processor.calculate_auc(
                self.dff_data,
                self.processed_time_points
            )

            self.show_status("Data processing completed", 3000)
            logger.info("Data processing completed successfully")
            logger.info(f"Processed data shape: {processed_data.shape}")
            logger.info(f"Time points shape: {self.processed_time_points.shape}")

        except Exception as e:
            error_msg = f"Error processing data: {str(e)}"
            self.show_status(error_msg, 5000)
            logger.error(error_msg)
            QMessageBox.critical(self, "Error", error_msg)

    def calculate_normalized_responses(self, group_name: str, well_ids: list) -> dict:
        """Calculate normalized responses for a group of wells"""
        if "ionomycin" in group_name.lower():
            return None

        ionomycin_responses = self.get_ionomycin_responses()
        if not ionomycin_responses:
            return None

        normalized_peaks = []
        for well_id in well_ids:
            well_idx = next(idx for idx in range(96) if self.well_data[idx]["well_id"] == well_id)
            sample_id = self.well_data[well_idx].get("sample_id", "default")
            ionomycin_response = ionomycin_responses.get(sample_id)

            if ionomycin_response:
                peak = self.dff_data.loc[well_id].max()
                normalized_peaks.append((peak / ionomycin_response) * 100)

        if normalized_peaks:
            normalized_peaks = np.array(normalized_peaks)
            return {
                'mean': float(np.mean(normalized_peaks)),
                'sem': float(np.std(normalized_peaks) / np.sqrt(len(normalized_peaks)))
            }

        return None

    def update_plots(self):
        """Update all plot windows"""
        if self.raw_data is None:
            return

        # Clear existing plots
        if self.raw_plot_window.isVisible():
            self.raw_plot_window.clear_plot()
        if self.dff_plot_window.isVisible():
            self.dff_plot_window.clear_plot()

        # Get appropriate time values
        if self.remove_artifact:
            times = self.processed_time_points
        else:
            times = pd.to_numeric(self.raw_data.columns, errors='coerce')

        # Update plots for each selected well
        for idx in self.selected_wells:
            well_id = self.well_data[idx]["well_id"]
            if well_id in self.raw_data.index:
                # Plot raw data
                if self.raw_plot_window.isVisible():
                    if self.remove_artifact:
                        # Get processed raw data
                        n_cols = self.raw_data.shape[1]
                        start_idx = int(n_cols * self.analysis_params['artifact_start']/220)
                        end_idx = int(n_cols * self.analysis_params['artifact_end']/220)
                        values = pd.concat([
                            self.raw_data.loc[well_id][:start_idx],
                            self.raw_data.loc[well_id][end_idx:]
                        ])
                    else:
                        values = self.raw_data.loc[well_id]
                    self.raw_plot_window.plot_trace(well_id, times, values, self.well_data[idx]["color"])

                # Plot ΔF/F₀ data
                if self.dff_plot_window.isVisible() and self.dff_data is not None:
                    values = self.dff_data.loc[well_id]
                    self.dff_plot_window.plot_trace(well_id, times, values, self.well_data[idx]["color"])


    def export_flipr_format(self, output_file):
        """Export plate layout to FLIPR .fmg format"""
        # Create plate data section
        plate_section = ["[CFLIPRPlateData]",
                        "Object=CFLIPRPlateData",
                        "TotalWells=96"]

        # Map wells to group IDs
        group_map = {}  # Keep track of unique groups
        next_group_id = 4  # Start after default groups (0-3)

        # Generate plate assignments
        for i in range(96):
            row = i // 12 + 1
            col = i % 12 + 1

            # Create unique group identifier
            well = self.well_data[i]
            if well["label"] or well["concentration"]:
                group_key = (well["label"], well["concentration"])
                if group_key not in group_map:
                    group_map[group_key] = next_group_id
                    next_group_id += 1
                group_id = group_map[group_key]
            else:
                group_id = 0  # Default/empty group

            plate_section.append(f"Row{row}Col{col}={group_id}")

        # Create groups section
        groups_section = [
            "[CFLIPRGroupArray]",
            f"Size={len(group_map) + 4}"  # Include default groups
        ]

        # Add default groups (0-3)
        default_groups = [
            ("NO_GROUP", 0),
            ("Positive Controls", 1),
            ("Negative Controls", 2),
            ("BF Controls", 3)
        ]

        for name, group_id in default_groups:
            groups_section.append(f"[CFLIPRGroup{group_id}]")
            groups_section.extend([
                "Object=CFLIPRGroup",
                f"GroupID={group_id}",
                "Concentration=10",
                "ConcentrationUnits=µM",
                "Color=16777215",
                f"Type={1 if group_id > 0 else 0}",
                "Operation=0",
                "StartValue=0",
                "IncrementValue=0.1",
                "Direction=0",
                "Replicate=2",
                "ReplicateCount=1",
                "CurrentIndex=1",
                f"GroupName={name}",
                "Notes=",
                "ExcludeFromStatisticChart=FALSE"
            ])

        # Add user-defined groups
        for (label, conc), group_id in group_map.items():
            groups_section.append(f"[CFLIPRGroup{group_id}]")

            # Find first well with this group to get color
            well_idx = next(i for i in range(96)
                           if self.well_data[i]["label"] == label
                           and self.well_data[i]["concentration"] == conc)
            color = rgb_to_decimal(self.well_data[well_idx]["color"])

            groups_section.extend([
                "Object=CFLIPRGroup",
                f"GroupID={group_id}",
                f"Concentration={format_concentration(conc)}",
                "ConcentrationUnits=µM",
                f"Color={color}",
                "Type=2",  # User-defined group
                "Operation=0",
                "StartValue=0",
                "IncrementValue=0.1",
                "Direction=0",
                "Replicate=2",
                "ReplicateCount=1",
                "CurrentIndex=1",
                f"GroupName={label}",
                "Notes=",
                "ExcludeFromStatisticChart=FALSE"
            ])

        # Write complete file
        with open(output_file, 'w') as f:
            f.write('\n'.join(plate_section + groups_section))

    def export_to_flipr(self):
        """Export current plate layout to FLIPR .fmg format"""
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export FLIPR Layout", "",
            "FLIPR Files (*.fmg);;All Files (*)",
            options=options
        )
        if file_path:
            try:
                self.export_flipr_format(file_path)
                QMessageBox.information(self, "Success",
                                      "Plate layout exported successfully")
            except Exception as e:
                QMessageBox.critical(self, "Error",
                                   f"Failed to export FLIPR layout: {str(e)}")



if __name__ == "__main__":
    app = QApplication(sys.argv)

    # Set the application style
    app.setStyle('Fusion')

    window = WellPlateLabeler()
    window.show()
    sys.exit(app.exec_())
